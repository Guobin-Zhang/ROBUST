import os
import time
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import cv2

class ExcelDataManager:
    def __init__(self, save_dir='./video_results/excel_data'):
        self.save_dir = save_dir
        os.makedirs(save_dir, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
    def create_training_curves_sheet(self, history, sheet_name='Training Curves'):
        ws = self.wb.create_sheet(title=sheet_name)
        headers = ['Iteration', 'Current PSNR', 'Previous PSNR', 'Current Acc', 'Previous Acc', 'AEAN Weight Norm']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        iterations = len(history['curr_psnr'])
        for i in range(iterations):
            row = [
                i,
                history['curr_psnr'][i] if i < len(history['curr_psnr']) else '',
                history['prev_psnr'][i] if i < len(history['prev_psnr']) else '',
                history['curr_acc'][i] if i < len(history['curr_acc']) else '',
                history['prev_acc'][i] if i < len(history['prev_acc']) else '',
                history['aean_weight_norm'][i] if 'aean_weight_norm' in history and i < len(history['aean_weight_norm']) else ''
            ]
            ws.append(row)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        return ws
    
    def create_frame_comparison_sheet(self, frame_metrics, sheet_name='Frame Comparison'):
        ws = self.wb.create_sheet(title=sheet_name)
        headers = ['Frame Index', 'PSNR (dB)', 'MSE', 'SSIM', 'Bit Accuracy']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for i, metrics in enumerate(frame_metrics):
            row = [
                i,
                metrics.get('psnr', 0),
                metrics.get('mse', 0),
                metrics.get('ssim', 0),
                metrics.get('bit_acc', 0)
            ]
            ws.append(row)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        return ws
    
    def create_video_metrics_sheet(self, video_metrics, sheet_name='Video Metrics'):
        ws = self.wb.create_sheet(title=sheet_name)
        headers = ['Frame', 'Original vs Stego PSNR', 'Original vs Stego MSE', 
                   'Tracking vs Stego PSNR', 'Tracking vs Stego MSE',
                   'Payload Accuracy', 'SSIM']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for i, metrics in enumerate(video_metrics):
            row = [
                i,
                metrics.get('orig_stego_psnr', 0),
                metrics.get('orig_stego_mse', 0),
                metrics.get('track_stego_psnr', 0),
                metrics.get('track_stego_mse', 0),
                metrics.get('payload_acc', 0),
                metrics.get('ssim', 0)
            ]
            ws.append(row)
        ws.append([])
        avg_row = ['Average',
                   np.mean([m['orig_stego_psnr'] for m in video_metrics]),
                   np.mean([m['orig_stego_mse'] for m in video_metrics]),
                   np.mean([m['track_stego_psnr'] for m in video_metrics]),
                   np.mean([m['track_stego_mse'] for m in video_metrics]),
                   np.mean([m['payload_acc'] for m in video_metrics]),
                   np.mean([m['ssim'] for m in video_metrics])]
        ws.append(avg_row)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        return ws
    
    def save_workbook(self, filename='video_stego_data.xlsx'):
        filepath = os.path.join(self.save_dir, filename)
        self.wb.save(filepath)
        print(f"Excel workbook saved: {filepath}")
        return filepath

class VideoStegoVisualizer:
    def __init__(self, save_dir='./video_results', create_timestamp_subdir=True):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        if create_timestamp_subdir:
            self.save_dir = os.path.join(save_dir, timestamp)
        else:
            self.save_dir = save_dir
        os.makedirs(self.save_dir, exist_ok=True)
        self.subdirs = {
            'training_curves': os.path.join(self.save_dir, 'training_curves'),
            'scatter_plots': os.path.join(self.save_dir, 'scatter_plots'),
            'frame_comparisons': os.path.join(self.save_dir, 'frame_comparisons'),
            'generated_videos': os.path.join(self.save_dir, 'generated_videos'),
            'tracking_visualizations': os.path.join(self.save_dir, 'tracking_visualizations'),
            'excel_data': os.path.join(self.save_dir, 'excel_data')
        }
        for d in self.subdirs.values():
            os.makedirs(d, exist_ok=True)
        self.excel_manager = ExcelDataManager(self.subdirs['excel_data'])
        print(f"Results will be saved to: {self.save_dir}")
    
    def plot_training_curves_with_prev_curr(self, history, save_name='training_curves.png', iteration=None):
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        iterations = range(len(history['curr_psnr']))
        if len(history['curr_psnr']) > 0:
            avg_psnr = np.mean(history['curr_psnr'])
            avg_acc = np.mean(history['curr_acc'])
        else:
            avg_psnr = 0
            avg_acc = 0
        axes[0, 0].plot(iterations, history['curr_psnr'], 'b-', linewidth=2, label='Current Frame')
        axes[0, 0].plot(iterations, history['prev_psnr'], 'r--', linewidth=1.5, alpha=0.7, label='Previous Frame')
        axes[0, 0].axhline(y=avg_psnr, color='g', linestyle='-.', linewidth=2, label=f'Avg ({avg_psnr:.2f})')
        axes[0, 0].set_xlabel('Frame Index')
        axes[0, 0].set_ylabel('PSNR (dB)')
        axes[0, 0].set_title('PSNR: Current vs Previous Frame')
        axes[0, 0].legend()
        axes[0, 0].grid(True, alpha=0.3)
        axes[0, 1].plot(iterations, history['curr_acc'], 'b-', linewidth=2, label='Current Frame')
        axes[0, 1].plot(iterations, history['prev_acc'], 'r--', linewidth=1.5, alpha=0.7, label='Previous Frame')
        axes[0, 1].axhline(y=avg_acc, color='g', linestyle='-.', linewidth=2, label=f'Avg ({avg_acc:.4f})')
        axes[0, 1].set_xlabel('Frame Index')
        axes[0, 1].set_ylabel('Accuracy')
        axes[0, 1].set_title('Accuracy: Current vs Previous Frame')
        axes[0, 1].legend()
        axes[0, 1].grid(True, alpha=0.3)
        if len(history['curr_psnr']) > 10:
            window = min(10, len(history['curr_psnr']) // 2)
            curr_psnr_smooth = np.convolve(history['curr_psnr'], np.ones(window)/window, mode='valid')
            prev_psnr_smooth = np.convolve(history['prev_psnr'], np.ones(window)/window, mode='valid')
            iter_smooth = range(len(curr_psnr_smooth))
            axes[1, 0].plot(iter_smooth, curr_psnr_smooth, 'b-', linewidth=2, label=f'Current (smooth)')
            axes[1, 0].plot(iter_smooth, prev_psnr_smooth, 'r--', linewidth=1.5, label=f'Previous (smooth)')
            axes[1, 0].set_xlabel('Frame Index')
            axes[1, 0].set_ylabel('PSNR (dB)')
            axes[1, 0].set_title('Smoothed PSNR Comparison')
            axes[1, 0].legend()
            axes[1, 0].grid(True, alpha=0.3)
        else:
            axes[1, 0].text(0.5, 0.5, 'Insufficient data for smoothing', 
                           ha='center', va='center', transform=axes[1, 0].transAxes)
            axes[1, 0].set_title('Smoothed PSNR')
        if 'aean_weight_norm' in history and len(history['aean_weight_norm']) > 0:
            axes[1, 1].plot(history['aean_weight_norm'], 'purple', linewidth=2)
            axes[1, 1].set_xlabel('Update Step')
            axes[1, 1].set_ylabel('Weight Norm')
            axes[1, 1].set_title('AEAN Weight Norm Evolution')
            axes[1, 1].grid(True, alpha=0.3)
        else:
            axes[1, 1].text(0.5, 0.5, 'AEAN tracking disabled', 
                           ha='center', va='center', transform=axes[1, 1].transAxes)
            axes[1, 1].set_title('AEAN Status')
        plt.tight_layout()
        save_path = os.path.join(self.subdirs['training_curves'], save_name)
        plt.savefig(save_path, dpi=200, bbox_inches='tight')
        plt.close()
        sheet_name = f'Training Curves Iter {iteration}' if iteration is not None else 'Training Curves'
        self.excel_manager.create_training_curves_sheet(history, sheet_name)
        print(f"Training curves saved: {save_path}")
    
    def visualize_frame_comparison(self, system, video_clip, original_width, original_height, iteration, save_name=None):
        if save_name is None:
            save_name = f'frame_comparison_iter_{iteration:04d}.png'
        system.encoder.eval()
        system.decoder.eval()
        frame_metrics = []
        with torch.no_grad():
            video_clip = video_clip.to(system.device)
            B, C, T, H, W = video_clip.shape
            fig, axes = plt.subplots(2, min(T, 4), figsize=(16, 8))
            for t in range(min(T, 4)):
                frame = video_clip[0, :, t, :, :].unsqueeze(0)
                people_info = system.people_tracker.detect_and_track_people(frame)
                tracking_regions = system.people_tracker.get_tracking_regions(frame, people_info)
                secret_tensor = system.people_tracker.extract_secret_from_overlay(people_info, frame)
                payload = torch.zeros(1, system.config.data_depth, H, W).to(system.device)
                payload_flat = payload.view(1, -1)
                secret_len = min(secret_tensor.numel(), payload_flat.numel())
                payload_flat[0, :secret_len] = secret_tensor[:secret_len].to(system.device)
                payload = payload_flat.view(1, system.config.data_depth, H, W)
                stego_final, mask = system.apply_stego_to_tracking_regions(frame, payload, tracking_regions, t)
                frame_with_tracking = system.people_tracker.draw_overlay(frame, people_info)
                tracking_np = ((frame_with_tracking[0].cpu().permute(1,2,0).numpy() + 1) / 2).clip(0, 1)
                stego_np = ((stego_final[0].cpu().permute(1,2,0).numpy() + 1) / 2).clip(0, 1)
                mse = np.mean((stego_np - tracking_np) ** 2)
                psnr = 10 * np.log10(1.0 / (mse + 1e-10))
                decoded = system.decoder(stego_final)
                decoded_binary = (torch.sigmoid(decoded) > 0.5).float()
                payload_binary = ((payload + 1) / 2 > 0.5).float()
                bit_acc = (decoded_binary == payload_binary).float().mean().item()
                ssim_val = system._calculate_ssim(frame_with_tracking, stego_final).item()
                frame_metrics.append({
                    'psnr': psnr,
                    'mse': mse,
                    'ssim': ssim_val,
                    'bit_acc': bit_acc
                })
                axes[0, t].imshow(tracking_np)
                axes[0, t].set_title(f'Frame {t} Tracking')
                axes[0, t].axis('off')
                axes[1, t].imshow(stego_np)
                axes[1, t].set_title(f'Frame {t} Stego (PSNR: {psnr:.1f}dB)')
                axes[1, t].axis('off')
            plt.suptitle(f'Video Stego Frame Comparison - Iteration {iteration}', fontsize=14, fontweight='bold')
            plt.tight_layout()
            save_path = os.path.join(self.subdirs['frame_comparisons'], save_name)
            plt.savefig(save_path, dpi=150, bbox_inches='tight')
            plt.close()
            sheet_name = f'Frame Comparison Iter {iteration}'
            self.excel_manager.create_frame_comparison_sheet(frame_metrics, sheet_name=sheet_name)
            print(f"Frame comparison saved: {save_path}")
    
    def visualize_video_comparison(self, system, video_clip, original_width, original_height, iteration, save_name=None):
        if save_name is None:
            save_name = f'video_comparison_iter_{iteration:04d}.png'
        system.encoder.eval()
        system.decoder.eval()
        video_metrics = []
        with torch.no_grad():
            video_clip = video_clip.to(system.device)
            B, C, T, H, W = video_clip.shape
            n_frames = min(T, 8)
            fig, axes = plt.subplots(4, n_frames, figsize=(24, 12))
            for t in range(n_frames):
                frame = video_clip[0, :, t, :, :].unsqueeze(0)
                people_info = system.people_tracker.detect_and_track_people(frame)
                tracking_regions = system.people_tracker.get_tracking_regions(frame, people_info)
                secret_tensor = system.people_tracker.extract_secret_from_overlay(people_info, frame)
                payload = torch.zeros(1, system.config.data_depth, H, W).to(system.device)
                payload_flat = payload.view(1, -1)
                secret_len = min(secret_tensor.numel(), payload_flat.numel())
                payload_flat[0, :secret_len] = secret_tensor[:secret_len].to(system.device)
                payload = payload_flat.view(1, system.config.data_depth, H, W)
                stego_final, mask = system.apply_stego_to_tracking_regions(frame, payload, tracking_regions, t)
                frame_np = ((frame[0].cpu().permute(1,2,0).numpy() + 1) / 2).clip(0, 1)
                frame_with_tracking = system.people_tracker.draw_overlay(frame, people_info)
                tracking_np = ((frame_with_tracking[0].cpu().permute(1,2,0).numpy() + 1) / 2).clip(0, 1)
                stego_np = ((stego_final[0].cpu().permute(1,2,0).numpy() + 1) / 2).clip(0, 1)
                diff_np = np.abs(stego_np - tracking_np)
                diff_enhanced = np.clip(diff_np * 10, 0, 1)
                orig_stego_mse = np.mean((stego_np - frame_np) ** 2)
                orig_stego_psnr = 10 * np.log10(1.0 / (orig_stego_mse + 1e-10))
                track_stego_mse = np.mean((stego_np - tracking_np) ** 2)
                track_stego_psnr = 10 * np.log10(1.0 / (track_stego_mse + 1e-10))
                decoded = system.decoder(stego_final)
                decoded_binary = (torch.sigmoid(decoded) > 0.5).float()
                payload_binary = ((payload + 1) / 2 > 0.5).float()
                payload_acc = (decoded_binary == payload_binary).float().mean().item()
                ssim_val = system._calculate_ssim(frame, stego_final).item()
                video_metrics.append({
                    'orig_stego_psnr': orig_stego_psnr,
                    'orig_stego_mse': orig_stego_mse,
                    'track_stego_psnr': track_stego_psnr,
                    'track_stego_mse': track_stego_mse,
                    'payload_acc': payload_acc,
                    'ssim': ssim_val
                })
                axes[0, t].imshow(frame_np)
                axes[0, t].set_title(f'Frame {t}')
                axes[0, t].axis('off')
                axes[1, t].imshow(tracking_np)
                axes[1, t].set_title(f'Tracking')
                axes[1, t].axis('off')
                axes[2, t].imshow(stego_np)
                axes[2, t].set_title(f'Stego {orig_stego_psnr:.1f}dB')
                axes[2, t].axis('off')
                axes[3, t].imshow(diff_enhanced)
                axes[3, t].set_title(f'Diff (x10)')
                axes[3, t].axis('off')
            axes[0, 0].set_ylabel('Original', fontsize=12, fontweight='bold')
            axes[1, 0].set_ylabel('Tracking', fontsize=12, fontweight='bold')
            axes[2, 0].set_ylabel('Stego', fontsize=12, fontweight='bold')
            axes[3, 0].set_ylabel('Difference', fontsize=12, fontweight='bold')
            plt.suptitle(f'Video Stego with People Tracking - Iteration {iteration}', fontsize=16, fontweight='bold')
            plt.tight_layout()
            save_path = os.path.join(self.subdirs['tracking_visualizations'], save_name)
            plt.savefig(save_path, dpi=150, bbox_inches='tight')
            plt.close()
            sheet_name = f'Video Metrics Iter {iteration}'
            self.excel_manager.create_video_metrics_sheet(video_metrics, sheet_name=sheet_name)
            print(f"Video comparison with tracking saved: {save_path}")
    
    def save_all_excel_data(self, filename='video_stego_complete_data.xlsx'):
        return self.excel_manager.save_workbook(filename)